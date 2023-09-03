from faker import Faker
from openpyxl import Workbook
# Faker Method and their	Purpose
# name()	It is used to generate a fake name.
# address()	It is used to generate a fake address.
# email()	It is used to generate fake email
# url()	It is used to generate a fake url address.
# phone_number()	It is used to generate a fake phone number.
# country()	It is used to generate a country name.
# text()	It is used to generate fake text.
# sentence()	It is used to generate large text.
# date()	It is used to generate a dummy date value.
# time()	It is used to generate a dummy time value.
# year()	It is used to generate a dummy year value.

fakeData = Faker(["hi_IN", "en_IN"])  # it take argument as fake data in different languages
wb = Workbook()
sheet= wb.active

# print(fakeData.name())
# print(fakeData.address())
# print(fakeData.email())
# print(fakeData.country())

for row in range(1,100):
    for cell in range(1,4):
        sheet.cell(row,1).value= fakeData.name()
        sheet.cell(row, 2).value = fakeData.email()
        sheet.cell(row, 3).value = fakeData.address()
        sheet.cell(row, 4).value = fakeData.phone_number()
wb.save("C:\\Users\\Lenovo\\Desktop\\Python\\excelData.xlsx")
