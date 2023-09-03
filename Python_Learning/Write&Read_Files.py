from openpyxl import Workbook, load_workbook

"""
Manual steps to write file
Open note pad and create file
Write in file
Close this file
"""

"""
Mode
Write mode: w [It overide date with new data on old dat]
Read mode: r
Write and Read mode: r++
Append mode: a  [Does not override it create new data]
"""

# f = open("C:\\Users\Lenovo\Desktop\Python\writeData.txt", "r" )
# print(f.read())
# f.close()

# =================Write and read data of .txt file========================
# with open("C:\\Users\Lenovo\Desktop\Python\writeData.txt", "r") as f:
#     print(f.read())
# with open("C:\\Users\Lenovo\Desktop\Python\writeData.txt", "a") as f:
#     print(f.write("7576878787"+"\n"))


# =================Write data of .xlsx file======================
# wb = Workbook()   # Create object of Workbookfactory class
# sheet= wb.active  # Focus on active sheet

# sheet["A1"] = "Pratik"  # Write into sheet
# wb.save("C:\\Users\\Lenovo\\Desktop\\Python\\excelData.xlsx")  # Save written data into excel file
#
# Dict = [["name","age"], ["Pratik","26"], ["Shubham","27"], ["Prajwal","22"]]
# for data in Dict:
#     sheet.append(data)  # Append data  into excel without overide
# wb.save("C:\\Users\\Lenovo\\Desktop\\Python\\excelData.xlsx")

# for row in range(1,5):
#     for cell in range(65, 68):
#         print(chr(cell)+str(row))
#         sheet[chr(cell)+str(row)]= "TestData"
# wb.save("C:\\Users\\Lenovo\\Desktop\\Python\\excelData.xlsx")
#
# for row in range(1,5):  # Row
#     for cell in range(1, 5): # Cell
#         sheet.cell(row=row, column=cell).value = "Prajwal"
# wb.save("C:\\Users\\Lenovo\\Desktop\\Python\\excelData.xlsx")

# =================Read data of .xlsx file======================

wb = load_workbook(filename="C:\\Users\\Lenovo\\Desktop\\Python\\excelData.xlsx")
sheet = wb["Sheet"]
print(sheet.cell(1,1).value)
print("Max no: ",sheet.max_row)
print("Max no: ",sheet.max_column)


